#!/usr/bin/env python3
"""Convert XLSX files to CSV format."""

import argparse
import csv
import sys

from openpyxl import load_workbook


def find_header_row(rows: list) -> int:
    """Find the header row index by looking for the first row where most cells are non-empty."""
    for i, row in enumerate(rows):
        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())
        if non_empty > 1:
            return i
    return 0


def xlsx_to_csv(input_file: str, output_file: str) -> None:
    """Read the first worksheet of an XLSX file and write it as CSV."""
    wb = load_workbook(input_file, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(rows)

    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        for row in rows[header_idx:]:
            writer.writerow(
                ['' if cell is None else cell for cell in row]
            )

    wb.close()


def main():
    parser = argparse.ArgumentParser(description="Convert XLSX to CSV.")
    parser.add_argument("-i", "--input", required=True, help="Input XLSX file path.")
    parser.add_argument("-o", "--output", required=True, help="Output CSV file path.")
    args = parser.parse_args()

    try:
        xlsx_to_csv(args.input, args.output)
        print(f"Successfully converted '{args.input}' to '{args.output}'")
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
